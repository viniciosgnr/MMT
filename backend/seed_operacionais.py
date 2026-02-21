"""Seed operational samples from the Operacionais sheet of the spreadsheet."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from datetime import datetime, date, timedelta
from app.database import SessionLocal
from app import models

SPREADSHEET = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "MMT Specifications", "06 - Sampling Plan (2).xlsm"
)

# FPSO short code → full name mapping  
FPSO_MAP = {
    "CDS": "CDS - Cidade de Saquarema",
    "CDM": "CDM - Cidade de Maricá",
    "CDI": "CDI - Cidade de Ilhabela",
    "CDP": "CDP - Cidade de Paraty",
    "ESS": "ESS - Espírito Santo",
    "CPX": "CPX - Capixaba",
    "CDA": "CDA - Cidade de Anchieta",
    "ADG": "ADG - Alexandre de Gusmão",
    "ATD": "ATD - Almirante Tamandaré",
    "SEP": "SEP - Sepetiba",
}

STATUS_MAP = {
    "Aprovado": "Report approved/reproved",
    "Reprovado": "Report approved/reproved",
}

VALIDATION_MAP = {
    "Aprovado": "Approved",
    "Reprovado": "Reproved",
}


def to_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
    except:
        return None


def seed():
    if not os.path.exists(SPREADSHEET):
        print(f"❌ Spreadsheet not found: {SPREADSHEET}")
        return

    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    ws = wb["Operacionais"]

    db = SessionLocal()
    created = 0
    skipped = 0

    try:
        for r in range(2, ws.max_row + 1):
            fpso_code = ws.cell(r, 2).value
            tag = ws.cell(r, 3).value
            tipo = ws.cell(r, 4).value
            status_xl = ws.cell(r, 17).value

            if not fpso_code or not tag:
                continue

            fpso_name = FPSO_MAP.get(str(fpso_code).strip(), str(fpso_code))

            # Find or create sample point (tag_number is unique globally)
            point = (
                db.query(models.SamplePoint)
                .filter(models.SamplePoint.tag_number == str(tag).strip())
                .first()
            )
            if not point:
                point = models.SamplePoint(
                    tag_number=str(tag).strip(),
                    fpso_name=fpso_name,
                    description=str(tipo).strip() if tipo else "Operacional",
                )
                db.add(point)
                db.flush()

            # Build sample ID
            sample_id = f"OP-{fpso_code}-{r:04d}"

            # Check if already exists
            if db.query(models.Sample).filter(models.Sample.sample_id == sample_id).first():
                skipped += 1
                continue

            # Parse dates
            sampling_date = to_date(ws.cell(r, 7).value)
            desemb_prev = to_date(ws.cell(r, 8).value)
            desemb_real = to_date(ws.cell(r, 9).value)
            lab_prev = to_date(ws.cell(r, 11).value)
            lab_real = to_date(ws.cell(r, 12).value)
            laudo_prev = to_date(ws.cell(r, 13).value)
            laudo_real = to_date(ws.cell(r, 14).value)
            cv_prev = to_date(ws.cell(r, 15).value)
            cv_real = to_date(ws.cell(r, 16).value)
            osm = ws.cell(r, 10).value
            laudo_n = ws.cell(r, 18).value

            # Determine status
            status = STATUS_MAP.get(str(status_xl).strip(), "Planned") if status_xl else "Planned"
            validation = VALIDATION_MAP.get(str(status_xl).strip()) if status_xl else None

            # If CV realizado exists, mark as Flow computer updated
            if cv_real:
                status = "Flow computer updated"

            # Calculate due_date based on status
            due_date = None
            if status == "Flow computer updated":
                due_date = cv_prev
            elif status == "Report approved/reproved":
                due_date = cv_prev  # next milestone is FC

            sample = models.Sample(
                sample_point_id=point.id,
                sample_id=sample_id,
                type=str(tipo).strip() if tipo else "Operacional",
                category="Operacional",
                status=status,
                responsible="Historical Import",
                osm_id=str(osm) if osm else None,
                laudo_number=str(laudo_n) if laudo_n else None,
                # Expected dates
                planned_date=sampling_date,
                disembark_expected_date=desemb_prev,
                lab_expected_date=lab_prev,
                report_expected_date=laudo_prev,
                fc_expected_date=cv_prev,
                # Actual dates
                sampling_date=sampling_date,
                disembark_date=desemb_real,
                delivery_date=lab_real,
                report_issue_date=laudo_real,
                fc_update_date=datetime.combine(cv_real, datetime.min.time()) if cv_real else None,
                due_date=due_date,
                validation_status=validation,
            )
            db.add(sample)
            created += 1

        db.commit()
        print(f"✅ Seed complete: {created} operational samples created, {skipped} skipped")
    finally:
        db.close()


if __name__ == "__main__":
    print("=" * 60)
    print("Seeding operational samples from spreadsheet...")
    print("=" * 60)
    seed()
